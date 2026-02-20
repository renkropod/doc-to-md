"""XLSX/CSV 파서 - openpyxl + pandas 기반"""

import csv
import logging
from pathlib import Path

from doc_to_md.exceptions import ParserError
from doc_to_md.parsers.base import BaseParser, ParseResult

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """XLSX/CSV 문서를 마크다운 테이블로 변환하는 파서"""

    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls", ".csv"]

    def parse(self, file_path: Path, **kwargs) -> ParseResult:
        """XLSX/CSV 파일을 마크다운 테이블로 변환"""
        suffix = file_path.suffix.lower()

        if suffix == ".csv":
            return self._parse_csv(file_path)

        try:
            return self._parse_with_openpyxl(file_path)
        except Exception as e:
            logger.warning("openpyxl 파싱 실패, pandas로 fallback: %s", e)

        try:
            return self._parse_with_pandas(file_path)
        except Exception as e:
            raise ParserError(f"XLSX 파싱 실패: {file_path} - {e}") from e

    def _parse_with_openpyxl(self, file_path: Path) -> ParseResult:
        """openpyxl을 사용한 XLSX 파싱"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ParserError(
                "XLSX 파싱을 위해 openpyxl을 설치해주세요: pip install openpyxl"
            )

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        parts: list[str] = []
        tables: list[str] = []
        metadata = {}

        if wb.properties:
            if wb.properties.title:
                metadata["title"] = wb.properties.title
            if wb.properties.creator:
                metadata["author"] = wb.properties.creator

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## {sheet_name}")

            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if any(cells):  # 빈 행 건너뛰기
                    rows.append(cells)

            if rows:
                md_table = self._rows_to_markdown(rows)
                tables.append(md_table)
                parts.append(md_table)

        wb.close()

        return ParseResult(
            content="\n\n".join(parts),
            metadata=metadata,
            tables=tables,
        )

    def _parse_with_pandas(self, file_path: Path) -> ParseResult:
        """pandas를 사용한 XLSX 파싱"""
        try:
            import pandas as pd
        except ImportError:
            raise ParserError(
                "XLSX 파싱을 위해 pandas를 설치해주세요: pip install pandas"
            )

        dfs = pd.read_excel(str(file_path), sheet_name=None)
        parts: list[str] = []
        tables: list[str] = []

        for sheet_name, df in dfs.items():
            parts.append(f"## {sheet_name}")
            md_table = df.to_markdown(index=False)
            if md_table:
                tables.append(md_table)
                parts.append(md_table)

        return ParseResult(
            content="\n\n".join(parts),
            tables=tables,
        )

    def _parse_csv(self, file_path: Path) -> ParseResult:
        """CSV 파일 파싱"""
        rows: list[list[str]] = []

        # 인코딩 자동 감지 시도
        for encoding in ["utf-8", "cp949", "euc-kr", "utf-8-sig"]:
            try:
                with open(file_path, "r", encoding=encoding, newline="") as f:
                    reader = csv.reader(f)
                    rows = [row for row in reader]
                break
            except (UnicodeDecodeError, UnicodeError):
                continue

        if not rows:
            raise ParserError(f"CSV 파일 인코딩을 감지할 수 없습니다: {file_path}")

        md_table = self._rows_to_markdown(rows)

        return ParseResult(
            content=md_table,
            tables=[md_table],
        )

    @staticmethod
    def _rows_to_markdown(rows: list[list[str]]) -> str:
        """2D 리스트를 마크다운 테이블로 변환"""
        if not rows:
            return ""

        col_count = max(len(r) for r in rows)

        def pad_row(row: list[str]) -> list[str]:
            padded = row + [""] * (col_count - len(row))
            return [c.replace("|", "\\|").replace("\n", " ") for c in padded[:col_count]]

        header = pad_row(rows[0])
        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join("---" for _ in range(col_count)) + " |",
        ]

        for row in rows[1:]:
            lines.append("| " + " | ".join(pad_row(row)) + " |")

        return "\n".join(lines)
